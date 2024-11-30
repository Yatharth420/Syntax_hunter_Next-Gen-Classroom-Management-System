import pandas as pd
import random
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Function to simulate incoming attendance data
def generate_random_data():
    names = [f"Student_{i}" for i in range(1, 21)]
    rolls = [f"R{i:03}" for i in range(1, 21)]
    current_time = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
    return {
        "Name": random.choice(names),
        "Roll": random.choice(rolls),
        "Time": current_time
    }

# Path to the Excel file
excel_path = 'Attendance/Attendance-{datetoday}.xlsx'

# Create or load the Excel file
try:
    workbook = load_workbook(excel_path)
    sheet = workbook.active
except FileNotFoundError:
    # Initialize Excel file if not present
    df = pd.DataFrame(columns=["Name", "Roll", "Time"])
    df.to_excel(excel_path, index=False)
    workbook = load_workbook(excel_path)
    sheet = workbook.active

# Real-time data update loop
for i in range(10):  # Simulate 10 updates
    # Generate random data
    data = generate_random_data()
    
    # Append data to the Excel sheet
    new_row = [data["Name"], data["Roll"], data["Time"]]
    sheet.append(new_row)
    workbook.save(excel_path)
    print(f"Added new data: {new_row}")
    
    # Pause for 1 second to simulate real-time updates
    time.sleep(1)

# Close the workbook
workbook.close()

# Load the updated Excel file
df = pd.read_excel('Attendance/Attendance-{datetoday}.xlsx')

# Visualization: Count of entries over time (grouped by hour)
df['Time'] = pd.to_datetime(df['Time'])
df['Hour'] = df['Time'].dt.hour
attendance_count_by_hour = df.groupby('Hour').size()

# Plot the updated attendance trend
plt.figure(figsize=(10, 6))
attendance_count_by_hour.plot(kind='bar', color='skyblue')
plt.title("Updated Attendance Trend by Hour", fontsize=14)
plt.xlabel("Hour of the Day", fontsize=12)
plt.ylabel("Number of Attendees", fontsize=12)
plt.grid(True)
plt.tight_layout()
plt.show()
